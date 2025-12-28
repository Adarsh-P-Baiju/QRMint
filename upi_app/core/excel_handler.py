import openpyxl
from openpyxl.drawing.image import Image as XLImage
import pandas as pd
from datetime import datetime
from .qr_generator import UPIQRGenerator

def generate_batch_excel_workbook(df):
    """
    Generates an openpyxl Workbook with embedded QR codes.
    Returns the workbook object.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batch QR Codes"
    
    headers = ["Date", "VPA", "Name", "Amount", "Note", "QR Code"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 35
    
    qr_gen = UPIQRGenerator()
    
    for index, row in df.iterrows():
        vpa = str(row['vpa'])
        name = str(row['name'])
        amount = str(row['amount'])
        note = str(row['note']) if not pd.isna(row['note']) else ""
        date_str = str(datetime.now().strftime("%Y-%m-%d %H:%M"))
        
        ws.append([date_str, vpa, name, amount, note])
        
        row_num = index + 2
        ws.row_dimensions[row_num].height = 160 
        
        try:
             pil_img = qr_gen.generate_qr(vpa, name, amount, note, logo_path=None)
             
             img_byte_arr = BytesIO()
             pil_img.save(img_byte_arr, format='PNG')
             img_byte_arr.seek(0)
             
             img = ExcelImage(img_byte_arr)
             img.width = 200
             img.height = 200
             
             ws.add_image(img, f"F{row_num}")
             
        except Exception as e:
            ws[f"F{row_num}"] = f"Error: {e}"
            
    return wb
